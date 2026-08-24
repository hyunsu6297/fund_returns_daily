"""One-time conversion of historical KFR price workbooks to API-shaped JSON."""

from __future__ import annotations

import gzip
import json
from datetime import date, datetime
from pathlib import Path

import openpyxl


BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = BASE_DIR / "data" / "kfr"
OUTPUT = OUTPUT_DIR / "prices_legacy_history.json.gz"
FIELDS = [
    "trade_day", "cm_seq", "composite_name", "fund_ksd_code", "invest_day", "fund_k_name",
    "ret", "price", "prev_price", "share_rate", "cul_ret", "kospi", "kosdaq", "sp500",
    "nasdaq", "kobi30", "kobi120", "econo_idx",
]


def value(item):
    if isinstance(item, (datetime, date)):
        return item.isoformat()[:10]
    return item


def main() -> None:
    by_key = {}
    files = sorted(path for path in BASE_DIR.glob("펀드 기준가*.xlsx") if not path.name.startswith("~$"))
    for path in files:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet = workbook["Data"] if "Data" in workbook.sheetnames else workbook.active
        try:
            for row in sheet.iter_rows(min_row=4, values_only=True):
                trade_day = value(row[1]) if len(row) > 1 else None
                fund_code = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
                if not trade_day or not fund_code:
                    continue
                values = [value(item) for item in row[1:18]]
                values.insert(1, None)
                record = dict(zip(FIELDS, values, strict=True))
                record["fund_ksd_code"] = fund_code
                by_key[(str(trade_day)[:10], fund_code)] = record
        finally:
            workbook.close()
    content = [by_key[key] for key in sorted(by_key)]
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    with gzip.open(OUTPUT, "wt", encoding="utf-8") as stream:
        json.dump({"content": content, "total_elements": len(content)}, stream, ensure_ascii=False, separators=(",", ":"))
    print(f"created {OUTPUT}: rows={len(content)}")


if __name__ == "__main__":
    main()
