"""Run K-FROMS fund price download with strict date-condition checks."""

from __future__ import annotations

import argparse
import zipfile
from datetime import date, datetime
from pathlib import Path

import download_fund_price as base


REQUESTED_DATES = {"start": None, "end": None}
ORIGINAL_CLICK_TEXT = base.click_text
ORIGINAL_OPEN_CONTEXT_MENU = base.open_context_menu_excel_download


def date_field(page, label_text: str):
    candidates = [
        f"xpath=//label[normalize-space()='{label_text}']/following::input[1]",
        f"xpath=//label[contains(normalize-space(), '{label_text}')]/following::input[1]",
        f"input[aria-label*='{label_text}']",
        f"input[placeholder*='{label_text}']",
    ]
    return base.first_visible(page, candidates)


def fill_date(page, label_text: str, value: str) -> None:
    field = date_field(page, label_text)
    field.click()
    field.press("Control+A")
    field.press("Backspace")
    field.type(value, delay=20)
    field.press("Tab")
    field.evaluate(
        """(node, value) => {
            const setter = Object.getOwnPropertyDescriptor(
                window.HTMLInputElement.prototype,
                'value'
            ).set;
            setter.call(node, value);
            node.dispatchEvent(new Event('input', { bubbles: true }));
            node.dispatchEvent(new Event('change', { bubbles: true }));
            node.dispatchEvent(new FocusEvent('blur', { bubbles: true }));
        }""",
        value,
    )
    actual = field.input_value(timeout=2500)
    if actual != value:
        raise RuntimeError(f"{label_text} 날짜 입력 실패: 기대={value}, 실제={actual}")
    if "시작" in label_text:
        REQUESTED_DATES["start"] = value
    if "종료" in label_text:
        REQUESTED_DATES["end"] = value


def assert_date_fields(page) -> None:
    expected_start = REQUESTED_DATES["start"]
    expected_end = REQUESTED_DATES["end"]
    actual_start = date_field(page, "시작일").input_value(timeout=2500)
    actual_end = date_field(page, "종료일").input_value(timeout=2500)
    if actual_start != expected_start or actual_end != expected_end:
        raise RuntimeError(
            "조회 전 날짜 조건 불일치: "
            f"시작일 기대={expected_start}, 실제={actual_start}; "
            f"종료일 기대={expected_end}, 실제={actual_end}"
        )


def click_text(page, text: str, timeout: int = 8000) -> None:
    if text == "검색":
        assert_date_fields(page)
    return ORIGINAL_CLICK_TEXT(page, text, timeout=timeout)


def earliest_trade_date_xlsx(path: Path) -> str | None:
    from openpyxl import load_workbook

    if not zipfile.is_zipfile(path):
        return None
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        header_row = None
        trade_col = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
            for col_idx, value in enumerate(row, start=1):
                if str(value).strip() == "기준일":
                    header_row = row_idx
                    trade_col = col_idx
                    break
            if trade_col:
                break
        if not trade_col:
            return None
        dates = []
        for (value,) in ws.iter_rows(
            min_row=(header_row or 1) + 1,
            min_col=trade_col,
            max_col=trade_col,
            values_only=True,
        ):
            if isinstance(value, datetime):
                dates.append(value.strftime("%Y-%m-%d"))
            elif isinstance(value, date):
                dates.append(value.strftime("%Y-%m-%d"))
            elif value:
                text = str(value).strip()[:10]
                try:
                    dates.append(base.normalize_date(text))
                except ValueError:
                    continue
        return min(dates) if dates else None
    finally:
        wb.close()


def open_context_menu_excel_download(page, download_dir: Path) -> Path:
    downloaded = ORIGINAL_OPEN_CONTEXT_MENU(page, download_dir)
    expected_start = REQUESTED_DATES["start"]
    actual = earliest_trade_date_xlsx(downloaded)
    if expected_start and actual and actual > expected_start:
        raise RuntimeError(
            "다운로드 파일 시작일이 요청 시작일보다 늦습니다: "
            f"요청={expected_start}, 파일 첫 기준일={actual}. "
            "사이트 기본 기간으로 조회된 것으로 보입니다."
        )
    return downloaded


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--start-date", default="2025-01-02")
    parser.add_argument("--end-date", default=base.previous_business_day().strftime("%Y-%m-%d"))
    parser.add_argument("--headed", action="store_true")
    args = parser.parse_args()

    base.fill_date = fill_date
    base.click_text = click_text
    base.open_context_menu_excel_download = open_context_menu_excel_download

    output = base.run(
        start_date=base.normalize_date(args.start_date),
        end_date=base.normalize_date(args.end_date),
        headless=not args.headed,
    )
    print(output)


if __name__ == "__main__":
    main()
