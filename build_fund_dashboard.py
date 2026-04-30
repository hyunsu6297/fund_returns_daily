import csv
import json
from collections import OrderedDict
from datetime import date, datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import openpyxl

BASE_DIR = Path(__file__).resolve().parent
SOURCE = BASE_DIR / "펀드 기준가.xlsx"
MAPPING = BASE_DIR / "mapping.xlsx"
MAPPING_CSV = BASE_DIR / "mapping.csv"
HANA_EMP = BASE_DIR / "Hana_EMP.xlsx"
PLOTLY_JS = BASE_DIR / "plotly-2.35.2.min.js"
KST = ZoneInfo("Asia/Seoul")
OUTPUT = BASE_DIR / f"fund_return_chart({datetime.now(KST):%y%m%d}).html"
WEB_OUTPUT = BASE_DIR / "docs" / "index.html"


def as_date_text(value):
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    return str(value)[:10]


def as_number(value):
    if value is None or value == "":
        return None
    try:
        return round(float(value), 6)
    except (TypeError, ValueError):
        return None


def mapping_rows():
    if MAPPING_CSV.exists():
        with MAPPING_CSV.open("r", encoding="utf-8-sig", newline="") as file:
            return list(csv.reader(file))
    workbook = openpyxl.load_workbook(MAPPING, read_only=True, data_only=True)
    sheet = workbook["Data"] if "Data" in workbook.sheetnames else workbook.active
    return list(sheet.iter_rows(values_only=True))


def load_mapping():
    rows = mapping_rows()
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    pos = {header: idx for idx, header in enumerate(headers)}
    mapping = OrderedDict()
    for row in rows[1:]:
        if not row:
            continue
        code = str(row[pos["Code"]]).strip() if "Code" in pos and row[pos["Code"]] is not None else ""
        name = str(row[pos["Fund"]]).strip() if "Fund" in pos and row[pos["Fund"]] is not None else ""
        group = str(row[pos["Group"]]).strip() if "Group" in pos and row[pos["Group"]] is not None else "기타"
        manager = str(row[pos["Manager"]]).strip() if "Manager" in pos and row[pos["Manager"]] is not None else "기타"
        team = str(row[pos["Team"]]).strip() if "Team" in pos and pos["Team"] < len(row) and row[pos["Team"]] is not None else ""
        if not code or not name or group == "BM":
            continue
        mapping[code] = {
            "code": code,
            "name": name,
            "type": group,
            "manager": manager,
            "team": team,
            "investDate": as_date_text(row[pos["LaunchDate"]]) if "LaunchDate" in pos and pos["LaunchDate"] < len(row) else "",
            "returnMode": "simple" if code.upper().startswith("EMP") else "compound",
        }
    return mapping


def append_hana_emp(mapping, funds, series, all_dates):
    if not HANA_EMP.exists():
        return
    workbook = openpyxl.load_workbook(HANA_EMP, read_only=True, data_only=True)
    sheet = workbook["일별"] if "일별" in workbook.sheetnames else workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = list(next(rows))
    emp_columns = [(idx, str(value).strip()) for idx, value in enumerate(headers[:5]) if value and str(value).strip() in mapping]
    prev_cum = {code: None for _, code in emp_columns}
    prev_nav = {code: None for _, code in emp_columns}
    for row in rows:
        base_date = as_date_text(row[0])
        if not base_date:
            continue
        for col_idx, code in emp_columns:
            cumulative_decimal = as_number(row[col_idx] if col_idx < len(row) else None)
            if cumulative_decimal is None:
                continue
            if code not in funds:
                funds[code] = len(funds)
                series[str(funds[code])] = []
            nav = 1000 * (1 + cumulative_decimal)
            daily = 0 if prev_cum[code] is None else (cumulative_decimal - prev_cum[code]) * 100
            series[str(funds[code])].append([
                base_date,
                round(daily, 6),
                round(nav, 6),
                round(prev_nav[code], 6) if prev_nav[code] is not None else None,
                0,
                round(cumulative_decimal * 100, 6),
                round(nav, 6),
            ])
            prev_cum[code] = cumulative_decimal
            prev_nav[code] = nav
            all_dates.append(base_date)


def build_payload():
    mapping = load_mapping()
    workbook = openpyxl.load_workbook(SOURCE, read_only=True, data_only=True)
    sheet = workbook["Data"] if "Data" in workbook.sheetnames else workbook.active
    funds = OrderedDict()
    series = {}
    bm_levels = OrderedDict()
    all_dates = []
    unmapped = {}

    for row in sheet.iter_rows(min_row=4, values_only=True):
        base_date = as_date_text(row[1])
        if not base_date:
            continue
        if base_date not in bm_levels:
            bm_levels[base_date] = {
                "BM_KOSPI": as_number(row[11]),
                "BM_KOSDAQ": as_number(row[12]),
                "BM_SPX": as_number(row[13]),
                "BM_NASDAQ": as_number(row[14]),
            }
        code = str(row[3]).strip() if row[3] is not None else ""
        info = mapping.get(code)
        if code and not info:
            item = unmapped.setdefault(code, {"latestDate": "", "type": "", "code": code, "investDate": "", "name": "", "dailyReturn": None, "nav": None, "cumulativeReturn": None, "count": 0})
            if base_date >= item["latestDate"]:
                item.update({
                    "latestDate": base_date,
                    "type": str(row[2]).strip() if row[2] is not None else "",
                    "investDate": as_date_text(row[4]),
                    "name": str(row[5]).strip() if row[5] is not None else "",
                    "dailyReturn": as_number(row[6]),
                    "nav": as_number(row[7]),
                    "cumulativeReturn": as_number(row[10]),
                })
            item["count"] += 1
        if not info:
            continue
        if code not in funds:
            funds[code] = len(funds)
            series[str(funds[code])] = []
        cum = as_number(row[10])
        level = round(cum * 10 + 1000, 6) if cum is not None else None
        series[str(funds[code])].append([base_date, as_number(row[6]), as_number(row[7]), as_number(row[8]), as_number(row[9]), cum, level])
        all_dates.append(base_date)

    append_hana_emp(mapping, funds, series, all_dates)

    fund_rows = []
    sorted_series = {}
    for code, old_id in funds.items():
        rows = sorted(series[str(old_id)], key=lambda item: item[0])
        fund_rows.append({"id": old_id, **mapping[code], "count": len(rows)})
    fund_rows.sort(key=lambda item: (item["name"], item["code"], item.get("investDate", "")))
    for new_id, fund in enumerate(fund_rows):
        old_id = fund["id"]
        fund = dict(fund)
        fund["id"] = new_id
        sorted_series[str(new_id)] = sorted(series[str(old_id)], key=lambda item: item[0])
        fund_rows[new_id] = fund

    bm_names = {"BM_KOSPI": "KOSPI", "BM_KOSDAQ": "KOSDAQ", "BM_SPX": "S&P500", "BM_NASDAQ": "NASDAQ"}
    bms = []
    for code, name in bm_names.items():
        rows = []
        first = prev = None
        for base_date, levels in sorted(bm_levels.items()):
            level = levels.get(code)
            if level is None:
                continue
            if first is None:
                first = level
            daily = 0 if prev is None else (level / prev - 1) * 100
            cum = (level / first - 1) * 100
            rows.append([base_date, round(daily, 6), round(level, 6), round(prev, 6) if prev is not None else None, 0, round(cum, 6), round(level, 6)])
            prev = level
        bms.append({"id": code, "name": name, "code": code, "type": "BM", "manager": "BM", "team": "BM", "investDate": "", "returnMode": "compound", "isBm": True, "count": len(rows)})
        sorted_series[code] = rows

    if not all_dates:
        all_dates = [date for rows in sorted_series.values() for date, *_ in rows]
    return {
        "sourceFile": SOURCE.name,
        "generatedAt": datetime.now(KST).strftime("%Y-%m-%d %H:%M"),
        "dateMin": min(all_dates),
        "dateMax": max(all_dates),
        "rowCount": sum(len(rows) for rows in sorted_series.values()),
        "fundCount": len(fund_rows),
        "funds": fund_rows,
        "bms": bms,
        "series": sorted_series,
        "unmappedFunds": sorted(unmapped.values(), key=lambda item: (item["latestDate"], item["name"], item["code"]), reverse=True),
    }


def build_html(payload):
    data = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    plotly = PLOTLY_JS.read_text(encoding="utf-8") if PLOTLY_JS.exists() else ""
    template = r'''<!doctype html>
<html lang="ko">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<meta http-equiv="Cache-Control" content="no-cache, no-store, must-revalidate">
<meta http-equiv="Pragma" content="no-cache">
<meta http-equiv="Expires" content="0">
<title>펀드 일별 수익률 대시보드</title>
<style>
:root{--ink:#17212b;--muted:#657385;--line:#d9e0e7;--panel:#fff;--wash:#f5f7fa;--accent:#0f766e;--good:#047857;--bad:#dc2626}*{box-sizing:border-box}body{margin:0;font-family:"Segoe UI","Malgun Gothic",Arial,sans-serif;color:var(--ink);background:var(--wash)}header{padding:26px 32px 18px;background:#fff;border-bottom:1px solid var(--line)}h1{margin:0 0 8px;font-size:26px}.source{color:var(--muted);font-size:13px;line-height:1.45}main{padding:22px 32px 32px}.shell{display:grid;grid-template-columns:320px minmax(0,1fr);gap:18px;align-items:start}.main{display:grid;gap:18px;min-width:0}.panel{background:var(--panel);border:1px solid var(--line);border-radius:8px}.sidebar{position:sticky;top:14px;max-height:calc(100vh - 28px);overflow:hidden;padding:14px}.side-title{margin:0 0 8px;font-size:14px}.bm-grid{display:grid;grid-template-columns:1fr 1fr;gap:6px;padding-bottom:10px;border-bottom:1px solid var(--line)}.bm-grid label{display:grid;grid-template-columns:18px 1fr;gap:6px;align-items:center;padding:7px 8px;border:1px solid #d8e0e8;border-radius:6px;background:#fff;font-size:13px;font-weight:700}.reset-all-row{display:grid;grid-template-columns:1fr 1fr;gap:8px;margin:10px 0 12px}.reset-all-row button{grid-column:1/-1}.team-actions{display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:5px;margin:8px 0 10px}.team-btn{width:100%;min-width:0;height:28px;padding:0 4px;color:#334155;font-size:11px;font-weight:750;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}.team-btn.active{color:#fff;background:var(--accent);border-color:var(--accent)}.filter-grid{display:grid;grid-template-columns:1fr 1fr;gap:8px;margin-top:8px}.filter{position:relative}.filter button,.action-row button,.period button,.chart-actions button,.table-actions button,.reset-all-row button,.team-btn,.legend-clear{border:1px solid #c8d1dc;border-radius:6px;background:#fff;color:var(--ink);font-weight:700;cursor:pointer}.filter>button{width:100%;height:38px;display:flex;justify-content:space-between;align-items:center;padding:0 10px}.filter>button:after{content:"▾";font-size:11px;color:var(--muted)}.menu{display:none;position:absolute;left:0;right:0;top:42px;z-index:20;max-height:420px;overflow:auto;background:#fff;border:1px solid #c8d1dc;border-radius:6px;box-shadow:0 12px 28px rgba(15,23,42,.14);padding:6px}.filter.open .menu{display:grid;gap:2px}.menu label{display:grid;grid-template-columns:18px 1fr;gap:7px;align-items:center;padding:7px 8px;border-radius:5px;font-size:13px}.menu label:hover{background:#f0f5f8}.action-row{display:grid;grid-template-columns:1fr 1fr;gap:8px;margin:8px 0}.action-row button,.reset-all-row button{height:38px}.fund-list{min-height:260px;max-height:calc(100vh - 292px);overflow:auto;border:1px solid #c8d1dc;border-radius:6px;background:#fff;padding:6px}.fund-option{display:grid;grid-template-columns:18px 1fr;gap:8px;padding:7px 8px;border-radius:5px;font-size:13px;font-weight:750;cursor:pointer}.fund-option:hover{background:#f0f5f8}.fund-meta{display:block;margin-top:2px;color:var(--muted);font-size:11px;font-weight:600}.period{padding:12px 14px;display:grid;grid-template-columns:minmax(220px,1fr) minmax(220px,1fr);gap:10px;align-items:end}.date-card{display:grid;gap:6px}.label-row{display:flex;gap:6px;align-items:center;color:#314052;font-size:12px;font-weight:750}.date-card input{height:38px;border:1px solid #c8d1dc;border-radius:6px;padding:0 10px;font-size:14px;font-weight:700}.quick{display:flex;gap:6px}.quick button{height:26px;padding:0 8px;font-size:12px}.metrics{display:grid;grid-template-columns:repeat(6,minmax(140px,1fr));overflow:hidden}.metric{padding:15px 16px;border-right:1px solid var(--line);min-height:86px}.metric:last-child{border-right:0}.metric span{display:block;color:var(--muted);font-size:12px;font-weight:700;margin-bottom:9px}.metric strong{font-size:22px;white-space:nowrap}.chart-panel{padding:16px}.panel-title{display:flex;justify-content:space-between;gap:12px;align-items:baseline;margin-bottom:10px}.title-note{color:var(--muted);font-size:11px}.chart-actions{display:flex;gap:10px;align-items:center;flex-wrap:wrap}.chart-actions button,.table-actions button,.legend-clear{height:26px;padding:0 9px;font-size:12px}h2{margin:0;font-size:17px}.title-with-note{display:flex;align-items:baseline;gap:10px;flex-wrap:wrap}#chart{height:468px;background:#fbfcfe;border:1px solid #e3e8ee;border-radius:6px}.chart-panel.collapsed #chart,.chart-panel.collapsed .legend,.chart-panel.collapsed .legend-head{display:none}.legend-head{display:flex;justify-content:flex-end;margin-top:10px}.legend{display:flex;flex-wrap:wrap;justify-content:center;gap:8px 14px;margin-top:8px}.legend button{display:inline-flex;gap:6px;align-items:center;border:1px solid transparent;border-radius:6px;background:transparent;padding:3px 6px;font-size:12px;font-weight:700;cursor:pointer}.legend button.active{background:#e8f3f1;border-color:var(--accent)}.legend button.dim{opacity:.28}.swatch{width:20px;height:3px;border-radius:2px}.table-panel{overflow:hidden}.table-head{display:flex;justify-content:space-between;align-items:baseline;padding:14px 16px;border-bottom:1px solid var(--line)}.table-wrap{max-height:430px;overflow:auto}.table-panel.expanded{position:relative;z-index:4;grid-column:1/-1}.table-panel.expanded .table-wrap{height:calc(100vh - 180px);min-height:760px;max-height:none!important}table{width:100%;border-collapse:collapse;background:#fff;font-size:13px}th,td{padding:9px 10px;border-bottom:1px solid #edf1f5;text-align:right;white-space:nowrap}th{position:sticky;top:0;background:#eef3f7;color:#314052;font-weight:800;cursor:pointer}th:first-child,td:first-child,th:nth-child(2),td:nth-child(2),th:nth-child(3),td:nth-child(3){text-align:left}.pos{color:var(--good);font-weight:750}.neg{color:var(--bad);font-weight:750}.empty{height:100%;min-height:320px;display:flex;align-items:center;justify-content:center;color:#5b6b82;font-size:16px}.hidden{display:none}@media(max-width:900px){.shell{grid-template-columns:1fr}.period{grid-template-columns:1fr}.metrics{grid-template-columns:1fr 1fr}.sidebar{position:relative;max-height:none}.fund-list{max-height:420px}}
</style>
<script>%%PLOTLY%%</script>
</head>
<body>
<header><h1>펀드 일별 수익률 대시보드</h1><div class="source">원본: %%SOURCE%% · 데이터 %%ROWS%%건 · 펀드 %%FUNDS%%개 · 기간 %%MIN%% ~ %%MAX%% · 생성 %%GEN%%</div></header>
<main><div class="shell"><aside class="panel sidebar"><h3 class="side-title">BM 선택</h3><div id="bmBox" class="bm-grid"></div><div class="reset-all-row"><button id="resetAll">전체 초기화</button></div><div id="teamButtons" class="team-actions"></div><h3 class="side-title">펀드 선택</h3><div class="filter-grid"><div class="filter" id="typeFilter"><div class="label-row">펀드 유형</div><button type="button" id="typeBtn">전체</button><div class="menu" id="typeMenu"></div></div><div class="filter" id="mgrFilter"><div class="label-row">운용사</div><button type="button" id="mgrBtn">전체</button><div class="menu" id="mgrMenu"></div></div></div><div class="action-row"><button id="selectBtn">일괄 선택</button><button id="clearBtn">일괄 해제</button></div><div id="fundBox" class="fund-list"></div></aside>
<div class="main"><section class="panel period"><div class="date-card"><div class="label-row"><span>시작일</span><span class="quick"><button id="ytd">연초</button><button id="mtd">월초</button></span></div><input type="date" id="start"></div><div class="date-card"><div class="label-row"><span>종료일</span><span class="quick"><button id="allPeriod">기간 전체</button></span></div><input type="date" id="end"></div></section><section class="panel metrics" id="metrics"></section><section class="panel chart-panel" id="chartPanel"><div class="panel-title"><div class="title-with-note"><h2>누적수익률 비교 추이</h2><span class="title-note">하단의 범례를 클릭하면 펀드가 강조됩니다</span></div><div class="chart-actions"><button id="toggleChart">차트 접기</button><button id="colors">색상 변경</button><button id="labels">레이블 숨김</button><span class="title-note" id="chartHint"></span></div></div><div id="chart"></div><div class="legend-head"><button id="clearHighlightBtn" class="legend-clear">강조해제</button></div><div id="legend" class="legend"></div></section><section class="panel table-panel" id="perfPanel"><div class="table-head"><div><h2>성과지표 <span class="title-note">(펀드 기준가로 계산한 성과로, 실제 집행금액 기준 성과와 일부 차이가 발생할 수 있습니다)</span></h2><div class="title-note" id="perfHint"></div></div><div class="table-actions"><button id="expandTable">확대</button><button id="downloadExcel">엑셀 다운</button></div></div><div class="table-wrap"><table><thead><tr id="perfHead"></tr></thead><tbody id="perfBody"></tbody></table></div></section><details class="panel table-panel"><summary class="table-head"><div><h2>미등록 펀드</h2><div class="title-note" id="unmappedHint"></div></div></summary><div class="table-wrap"><table><thead><tr><th>최근 기준일</th><th>유형</th><th>예탁원펀드코드</th><th>투자일</th><th>펀드명</th><th>일수익률</th><th>기준가</th><th>누적수익률</th><th>관측치</th></tr></thead><tbody id="unmappedBody"></tbody></table></div></details></div></div></main>
<script>
const DATA=%%DATA%%;
const typeOrder=['주식형','혼합형','메자닌','멀티','롱숏','공모주','EMP','Pre-IPO','채권형(원화)','채권형(외화)','기타'];
const colorPalettes=[['#2563eb','#dc2626','#059669','#d97706','#7c3aed','#0891b2','#be185d','#4d7c0f','#9333ea','#b45309','#0f766e','#e11d48'],['#111827','#f97316','#14b8a6','#c026d3','#84cc16','#ef4444','#0284c7','#a16207','#16a34a','#db2777','#4338ca','#65a30d'],['#0f766e','#e11d48','#1d4ed8','#ca8a04','#9333ea','#15803d','#ea580c','#0e7490','#be123c','#4f46e5','#854d0e','#047857']];
let selected=new Set(),selectedBm=new Set(),selectedTypes=new Set(),selectedManagers=new Set(),selectedTeams=new Set(),highlighted=new Set(),palette=0,showLabels=true,sortKey='periodReturn',sortDir='desc',currentModels=[];
const $=id=>document.getElementById(id);
const cols=[['name','펀드명'],['type','유형'],['investDate','투자일'],['periodReturn','기간수익률'],['annualReturn','연환산 수익률'],['vol','연환산 변동성'],['sharpe','Sharpe'],['mdd','MDD'],['ytd','YTD'],['mtd','MTD'],['d1','1일'],['w1','1주'],['m1','1개월'],['m3','3개월'],['m6','6개월'],['y1','1년']];
function esc(s){return String(s??'').replace(/[&<>"']/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c]))}
function fmtPct(v){return v==null||Number.isNaN(v)?'-':(v*100).toFixed(2)+'%'}function fmtPoint(v){return v==null||Number.isNaN(v)?'-':Number(v).toFixed(2)+'%'}function fmtNum(v,d=2){return v==null||Number.isNaN(v)?'-':Number(v).toLocaleString('ko-KR',{minimumFractionDigits:d,maximumFractionDigits:d})}function cls(v){return v>0?'pos':v<0?'neg':''}
function allDataDates(){return [...new Set(Object.values(DATA.series).flatMap(rows=>rows.map(row=>row[0])))].filter(Boolean).sort()}function firstDateOnOrAfter(target,end){return allDataDates().find(date=>date>=target&&date<=end)||target}
function dateRange(){let s=$('start').value,e=$('end').value;if(s>e)[s,e]=[e,s];return{start:s,end:e}}
function rowsFor(id,s=dateRange().start,e=dateRange().end){return(DATA.series[String(id)]||[]).filter(row=>row[0]>=s&&row[0]<=e)}
function stats(rows,fund){const lr=rows.filter(row=>row[6]!=null);if(lr.length<2)return{};const simple=fund.returnMode==='simple',first=lr[0][6],last=lr.at(-1)[6];const periodReturn=simple?(last-first)/1000:last/first-1;const rets=lr.slice(1).map((row,i)=>simple?(row[6]-lr[i][6])/1000:row[6]/lr[i][6]-1);const avg=rets.reduce((a,b)=>a+b,0)/rets.length;const variance=rets.length>1?rets.reduce((sum,v)=>sum+(v-avg)**2,0)/(rets.length-1):0;const vol=Math.sqrt(variance)*Math.sqrt(252);const annualReturn=simple?periodReturn*252/rets.length:Math.pow(1+periodReturn,252/rets.length)-1;let peak=1,mdd=0;for(const row of lr){const nav=simple?1+(row[6]-first)/1000:row[6]/first;peak=Math.max(peak,nav);mdd=Math.min(mdd,nav/peak-1)}return{observations:lr.length,periodReturn,annualReturn,vol,sharpe:vol?annualReturn/vol:null,mdd}}
function shiftDate(d,t,n){let x=new Date(d+'T00:00:00');if(t==='d')x.setDate(x.getDate()-n);if(t==='m')x.setMonth(x.getMonth()-n);if(t==='y')x.setFullYear(x.getFullYear()-n);return x.toISOString().slice(0,10)}
function trailing(fund){const end=dateRange().end,d=new Date(end+'T00:00:00');const ranges={ytd:`${d.getFullYear()}-01-01`,mtd:`${d.getFullYear()}-${String(d.getMonth()+1).padStart(2,'0')}-01`,d1:shiftDate(end,'d',1),w1:shiftDate(end,'d',7),m1:shiftDate(end,'m',1),m3:shiftDate(end,'m',3),m6:shiftDate(end,'m',6),y1:shiftDate(end,'y',1)};return Object.fromEntries(Object.entries(ranges).map(([k,s])=>[k,stats(rowsFor(fund.id,s<DATA.dateMin?DATA.dateMin:s,end),fund).periodReturn]))}
function teamValues(){return [...new Set(DATA.funds.map(f=>f.team).filter(Boolean))].sort((a,b)=>a.localeCompare(b,'ko-KR'))}
function baseByTeam(){return DATA.funds.filter(f=>!selectedTeams.size||selectedTeams.has(f.team))}
function filteredFunds(){return baseByTeam().filter(f=>(!selectedTypes.size||selectedTypes.has(f.type))&&(!selectedManagers.size||selectedManagers.has(f.manager)))}
function setAllFilters(){selectedTypes=new Set([...new Set(DATA.funds.map(f=>f.type))]);selectedManagers=new Set([...new Set(DATA.funds.map(f=>f.manager))])}
function resetFilterSelections(){selectedTypes.clear();selectedManagers.clear()}
function optionList(values,set,menu,btn){const visible=new Set([...set].filter(v=>values.includes(v)));const all=values.length>0&&visible.size===values.length;menu.innerHTML=`<label><input type="checkbox" data-all="1" ${all?'checked':''}>전체</label>`+values.map(v=>`<label><input type="checkbox" value="${esc(v)}" ${(all||visible.has(v))?'checked':''}>${esc(v)}</label>`).join('');btn.textContent=all?'전체':visible.size?`${visible.size}개 선택`:'선택 없음';menu.querySelectorAll('input').forEach(input=>input.onchange=()=>{if(input.dataset.all){if(all)values.forEach(v=>set.delete(v));else values.forEach(v=>set.add(v))}else{input.checked?set.add(input.value):set.delete(input.value)}renderFilters();renderFunds();render()})}
function renderFilters(){const base=baseByTeam();const types=[...new Set(base.map(f=>f.type))].sort((a,b)=>(typeOrder.indexOf(a)<0?99:typeOrder.indexOf(a))-(typeOrder.indexOf(b)<0?99:typeOrder.indexOf(b))||a.localeCompare(b,'ko-KR'));const mgrBase=base.filter(f=>!selectedTypes.size||selectedTypes.has(f.type));const managers=[...new Set(mgrBase.map(f=>f.manager))].sort((a,b)=>a.localeCompare(b,'ko-KR'));optionList(types,selectedTypes,$('typeMenu'),$('typeBtn'));optionList(managers,selectedManagers,$('mgrMenu'),$('mgrBtn'))}
function renderTeamButtons(){const teams=teamValues();$('teamButtons').innerHTML=teams.map(team=>`<button type="button" class="team-btn${selectedTeams.has(team)?' active':''}" data-team="${esc(team)}" title="${esc(team)}">${esc(team)}</button>`).join('');$('teamButtons').querySelectorAll('button').forEach(button=>button.onclick=()=>{const team=button.dataset.team;selectedTeams.has(team)?selectedTeams.delete(team):selectedTeams.add(team);resetFilterSelections();renderTeamButtons();renderFilters();renderFunds();render()})}
function renderFunds(){$('fundBox').innerHTML=filteredFunds().map(f=>`<label class="fund-option"><input type="checkbox" value="${f.id}" ${selected.has(String(f.id))?'checked':''}><span>${esc(f.name)}<span class="fund-meta">${esc([f.type,f.manager,f.team,f.code].filter(Boolean).join(' · '))}</span></span></label>`).join('');$('fundBox').querySelectorAll('input').forEach(input=>input.onchange=()=>{input.checked?selected.add(input.value):selected.delete(input.value);render()})}
function selectedItems(){return[...DATA.funds.filter(f=>selected.has(String(f.id))),...DATA.bms.filter(b=>selectedBm.has(String(b.id)))]}
function modelFor(fund,index){const rows=rowsFor(fund.id),color=colorPalettes[palette][index%colorPalettes[palette].length];return{fund,rows,color,stats:stats(rows,fund),periodReturns:trailing(fund)}}
function render(){currentModels=selectedItems().map(modelFor);renderMetrics();renderChart();renderPerformance();renderUnmapped()}
function renderMetrics(){const v=currentModels.filter(m=>m.stats.observations);const avg=k=>{const a=v.map(m=>m.stats[k]).filter(x=>x!=null&&!Number.isNaN(x));return a.length?a.reduce((s,x)=>s+x,0)/a.length:null};const best=[...v].sort((a,b)=>(b.stats.periodReturn??-9)-(a.stats.periodReturn??-9))[0];const bestS=[...v].sort((a,b)=>(b.stats.sharpe??-9)-(a.stats.sharpe??-9))[0];const worst=[...v].sort((a,b)=>(a.stats.mdd??9)-(b.stats.mdd??9))[0];const cards=[['선택 펀드',currentModels.length+'개'],['최고 기간수익률',best?fmtPct(best.stats.periodReturn):'-',best?.stats.periodReturn],['평균 연환산 수익률',fmtPct(avg('annualReturn')),avg('annualReturn')],['평균 연환산 변동성',fmtPct(avg('vol')),avg('vol')],['최고 Sharpe',bestS?.stats.sharpe==null?'-':fmtNum(bestS.stats.sharpe),bestS?.stats.sharpe],['최대 MDD',worst?fmtPct(worst.stats.mdd):'-',worst?.stats.mdd]];$('metrics').innerHTML=cards.map(c=>`<div class="metric"><span>${c[0]}</span><strong class="${cls(c[2])}">${c[1]}</strong></div>`).join('')}
function chartPoints(m){const lr=m.rows.filter(row=>row[6]!=null);if(lr.length<2)return[];const base=lr[0][6];return lr.map(row=>[row[0],m.fund.returnMode==='simple'?(row[6]-base)/1000*100:(row[6]/base-1)*100])}
function renderChart(){const series=currentModels.map(m=>({...m,points:chartPoints(m)})).filter(m=>m.points.length>=2);const validIds=new Set(series.map(m=>String(m.fund.id)));for(const id of [...highlighted])if(!validIds.has(id))highlighted.delete(id);$('chartHint').textContent=series.length?`시작점 0 기준 · ${series.length}개 펀드 비교`:'선택 기간에 누적수익률 데이터가 없습니다.';if(!series.length){if(window.Plotly)Plotly.purge('chart');$('chart').innerHTML='<div class="empty">1개 이상의 펀드를 선택하세요</div>';$('legend').innerHTML='';return}const traces=series.map(m=>{const on=highlighted.has(String(m.fund.id)),dim=highlighted.size&&!on;return{x:m.points.map(p=>p[0]),y:m.points.map(p=>p[1]),type:'scatter',mode:'lines',name:m.fund.name,opacity:dim?0.18:1,line:{color:m.color,width:on?(m.fund.isBm?6:4.2):(m.fund.isBm?4.2:2.4),dash:'solid'},hovertemplate:`<b>${esc(m.fund.name)}</b><br>기준일=%{x}<br>수익률=%{y:.2f}%<extra></extra>`}});const annotations=showLabels?series.map(m=>{const p=m.points.at(-1),on=highlighted.has(String(m.fund.id)),has=highlighted.size>0;return{order:on?1:0,a:{x:p[0],y:p[1],xref:'x',yref:'y',text:`${m.fund.name.slice(0,12)} ${p[1].toFixed(2)}%`,showarrow:true,arrowhead:0,arrowcolor:'rgba(0,0,0,0)',ax:24,ay:0,xanchor:'left',font:{size:on?13:11,color:on?'#111827':m.color},bgcolor:on?'rgba(255,255,255,.98)':has?'rgba(255,255,255,.58)':'rgba(255,255,255,.78)',bordercolor:on?m.color:has?'rgba(148,163,184,.45)':m.color,borderwidth:on?2:1,opacity:has&&!on?0.42:1}}}).sort((a,b)=>a.order-b.order).map(x=>x.a):[];Plotly.react('chart',traces,{margin:{l:58,r:28,t:18,b:42},paper_bgcolor:'#fbfcfe',plot_bgcolor:'#fbfcfe',hovermode:'closest',showlegend:false,annotations,xaxis:{showgrid:false,zeroline:false},yaxis:{ticksuffix:'%',gridcolor:'#e5eaf0',zerolinecolor:'#94a3b8'}},{responsive:true,displayModeBar:false});$('legend').innerHTML=series.map(m=>{const id=String(m.fund.id),active=highlighted.has(id),dim=highlighted.size&&!active;return`<button class="${active?'active ':''}${dim?'dim':''}" data-id="${esc(id)}"><span class="swatch" style="background:${m.color};${m.fund.isBm?'height:5px':''}"></span>${esc(m.fund.name.slice(0,42))}</button>`}).join('');$('legend').querySelectorAll('button').forEach(b=>b.onclick=()=>{highlighted.has(b.dataset.id)?highlighted.delete(b.dataset.id):highlighted.add(b.dataset.id);renderChart()})}
function perfModels(){const items=selectedItems();return(items.length?items:DATA.funds).map(modelFor)}function val(m,k){if(k==='name')return m.fund.name;if(k==='type')return m.fund.type||'';if(k==='investDate')return m.fund.investDate||'';if(['periodReturn','annualReturn','vol','sharpe','mdd'].includes(k))return m.stats[k];return m.periodReturns[k]}
function renderPerformance(){let ms=perfModels();$('perfHint').textContent=`${ms.length}개 펀드${selectedItems().length?'':' · 선택 없음: 전체 표시'}`;ms.sort((a,b)=>{const av=val(a,sortKey),bv=val(b,sortKey),dir=sortDir==='asc'?1:-1;if(typeof av==='string'||typeof bv==='string')return String(av??'').localeCompare(String(bv??''),'ko-KR')*dir;return((av??(sortDir==='asc'?Infinity:-Infinity))-(bv??(sortDir==='asc'?Infinity:-Infinity)))*dir});$('perfBody').innerHTML=ms.map(m=>`<tr><td>${esc(m.fund.name)}</td><td>${esc(m.fund.type||'')}</td><td>${esc(m.fund.investDate||'')}</td><td class="${cls(m.stats.periodReturn)}">${fmtPct(m.stats.periodReturn)}</td><td class="${cls(m.stats.annualReturn)}">${fmtPct(m.stats.annualReturn)}</td><td>${fmtPct(m.stats.vol)}</td><td class="${cls(m.stats.sharpe)}">${m.stats.sharpe==null?'-':fmtNum(m.stats.sharpe)}</td><td class="${cls(m.stats.mdd)}">${fmtPct(m.stats.mdd)}</td><td class="${cls(m.periodReturns.ytd)}">${fmtPct(m.periodReturns.ytd)}</td><td class="${cls(m.periodReturns.mtd)}">${fmtPct(m.periodReturns.mtd)}</td><td class="${cls(m.periodReturns.d1)}">${fmtPct(m.periodReturns.d1)}</td><td class="${cls(m.periodReturns.w1)}">${fmtPct(m.periodReturns.w1)}</td><td class="${cls(m.periodReturns.m1)}">${fmtPct(m.periodReturns.m1)}</td><td class="${cls(m.periodReturns.m3)}">${fmtPct(m.periodReturns.m3)}</td><td class="${cls(m.periodReturns.m6)}">${fmtPct(m.periodReturns.m6)}</td><td class="${cls(m.periodReturns.y1)}">${fmtPct(m.periodReturns.y1)}</td></tr>`).join('')}
function xmlEscape(v){return String(v??'').replace(/[<>&"']/g,ch=>({'<':'&lt;','>':'&gt;','&':'&amp;','"':'&quot;',"'":'&apos;'}[ch]))}function excelXml(sheet,header,rows){const trs=[header,...rows].map(row=>`<Row>${row.map(v=>`<Cell><Data ss:Type="String">${xmlEscape(v)}</Data></Cell>`).join('')}</Row>`).join('');return`<?xml version="1.0" encoding="UTF-8"?><?mso-application progid="Excel.Sheet"?><Workbook xmlns="urn:schemas-microsoft-com:office:spreadsheet" xmlns:ss="urn:schemas-microsoft-com:office:spreadsheet"><Worksheet ss:Name="${xmlEscape(sheet)}"><Table>${trs}</Table></Worksheet></Workbook>`}
function downloadExcel(){const header=cols.map(c=>c[1]);const rows=[...document.querySelectorAll('#perfBody tr')].map(tr=>[...tr.children].map(td=>td.textContent));const blob=new Blob(['\ufeff'+excelXml('성과지표',header,rows)],{type:'application/vnd.ms-excel;charset=utf-8'});const a=document.createElement('a');a.href=URL.createObjectURL(blob);a.download='performance_metrics.xls';a.click();URL.revokeObjectURL(a.href)}
function renderUnmapped(){const rows=DATA.unmappedFunds||[];$('unmappedHint').textContent=rows.length?`${rows.length.toLocaleString('ko-KR')}개 · mapping.xlsx 미등록`:'미등록 펀드 없음';$('unmappedBody').innerHTML=rows.length?rows.map(x=>`<tr><td>${esc(x.latestDate||'')}</td><td>${esc(x.type||'')}</td><td>${esc(x.code||'')}</td><td>${esc(x.investDate||'')}</td><td>${esc(x.name||'')}</td><td class="${cls(x.dailyReturn)}">${fmtPoint(x.dailyReturn)}</td><td>${fmtNum(x.nav)}</td><td class="${cls(x.cumulativeReturn)}">${fmtPoint(x.cumulativeReturn)}</td><td>${(x.count||0).toLocaleString('ko-KR')}</td></tr>`).join(''):`<tr><td colspan="9" class="empty">mapping.xlsx에 미등록된 펀드가 없습니다.</td></tr>`}
function resetAll(){selected.clear();selectedBm.clear();selectedTeams.clear();highlighted.clear();palette=0;showLabels=true;sortKey='periodReturn';sortDir='desc';$('start').value=DATA.dateMin;$('end').value=DATA.dateMax;$('labels').textContent='레이블 숨김';$('chartPanel').classList.remove('collapsed');$('toggleChart').textContent='차트 접기';setAllFilters();renderTeamButtons();renderFilters();renderFunds();$('bmBox').querySelectorAll('input').forEach(i=>i.checked=false);render()}
function init(){setAllFilters();$('start').value=DATA.dateMin;$('end').value=DATA.dateMax;$('start').min=DATA.dateMin;$('end').max=DATA.dateMax;$('bmBox').innerHTML=DATA.bms.map(b=>`<label><input type="checkbox" value="${b.id}">${esc(b.name)}</label>`).join('');$('bmBox').querySelectorAll('input').forEach(i=>i.onchange=()=>{i.checked?selectedBm.add(i.value):selectedBm.delete(i.value);render()});['typeFilter','mgrFilter'].forEach(id=>$(id).querySelector('button').onclick=()=>$(id).classList.toggle('open'));document.addEventListener('click',e=>{if(!e.target.closest('.filter'))document.querySelectorAll('.filter').forEach(f=>f.classList.remove('open'))});$('selectBtn').onclick=()=>{filteredFunds().forEach(f=>selected.add(String(f.id)));renderFunds();render()};$('clearBtn').onclick=()=>{filteredFunds().forEach(f=>selected.delete(String(f.id)));renderFunds();render()};$('resetAll').onclick=resetAll;$('ytd').onclick=()=>{const y=$('end').value.slice(0,4),raw=`${y}-01-01`,target=raw<DATA.dateMin?DATA.dateMin:raw;$('start').value=firstDateOnOrAfter(target,$('end').value);render()};$('mtd').onclick=()=>{const raw=$('end').value.slice(0,8)+'01',target=raw<DATA.dateMin?DATA.dateMin:raw;$('start').value=firstDateOnOrAfter(target,$('end').value);render()};$('allPeriod').onclick=()=>{$('start').value=DATA.dateMin;$('end').value=DATA.dateMax;render()};$('labels').onclick=()=>{showLabels=!showLabels;$('labels').textContent=showLabels?'레이블 숨김':'레이블 표시';render()};$('colors').onclick=()=>{palette=(palette+1)%colorPalettes.length;render()};$('toggleChart').onclick=()=>{$('chartPanel').classList.toggle('collapsed');$('toggleChart').textContent=$('chartPanel').classList.contains('collapsed')?'차트 열기':'차트 접기'};$('clearHighlightBtn').onclick=()=>{highlighted.clear();renderChart()};$('start').onchange=render;$('end').onchange=render;$('expandTable').onclick=()=>{$('perfPanel').classList.toggle('expanded');$('expandTable').textContent=$('perfPanel').classList.contains('expanded')?'축소':'확대'};$('downloadExcel').onclick=downloadExcel;$('perfHead').innerHTML=cols.map(c=>`<th data-key="${c[0]}">${c[1]}${sortKey===c[0]?(sortDir==='desc'?' ▼':' ▲'):''}</th>`).join('');$('perfHead').querySelectorAll('th').forEach(th=>th.onclick=()=>{sortDir=sortKey===th.dataset.key&&sortDir==='desc'?'asc':'desc';sortKey=th.dataset.key;$('perfHead').innerHTML=cols.map(c=>`<th data-key="${c[0]}">${c[1]}${sortKey===c[0]?(sortDir==='desc'?' ▼':' ▲'):''}</th>`).join('');$('perfHead').querySelectorAll('th').forEach(h=>h.onclick=th.onclick);renderPerformance()});renderTeamButtons();renderFilters();renderFunds();renderUnmapped();render()}
function startAutoRefreshCheck(){if(!/^https?:/.test(location.protocol))return;const current=DATA.generatedAt;const check=()=>fetch(location.href.split('#')[0]+(location.href.includes('?')?'&':'?')+'_='+Date.now(),{cache:'no-store'}).then(r=>r.ok?r.text():'').then(t=>{const m='"generatedAt":"',i=t.indexOf(m);if(i<0)return;const latest=t.slice(i+m.length).split('"')[0];if(latest&&latest!==current)location.reload()}).catch(()=>{});setInterval(check,300000);document.addEventListener('visibilitychange',()=>{if(!document.hidden)check()})}
init();startAutoRefreshCheck();
</script>
</body>
</html>'''
    return (template.replace("%%PLOTLY%%", plotly)
        .replace("%%DATA%%", data)
        .replace("%%SOURCE%%", payload["sourceFile"])
        .replace("%%ROWS%%", f"{payload['rowCount']:,}")
        .replace("%%FUNDS%%", f"{payload['fundCount']:,}")
        .replace("%%MIN%%", payload["dateMin"])
        .replace("%%MAX%%", payload["dateMax"])
        .replace("%%GEN%%", payload["generatedAt"]))


if __name__ == "__main__":
    html = build_html(build_payload())
    OUTPUT.write_text(html, encoding="utf-8")
    WEB_OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    WEB_OUTPUT.write_text(html, encoding="utf-8")
    print(OUTPUT)
    print(WEB_OUTPUT)
